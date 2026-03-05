import logging
import subprocess
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

logger = logging.getLogger(__name__)

# 시트 텍스트가 이 글자 수를 넘으면 행 단위로 분할
MAX_CHUNK_CHARS = 4000


def _convert_xls_to_xlsx(xls_path: Path) -> Path:
    """LibreOffice로 .xls → .xlsx 변환."""
    result = subprocess.run(
        [
            "libreoffice",
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(xls_path.parent),
            str(xls_path),
        ],
        capture_output=True,
        text=True,
        timeout=120,
    )

    xlsx_path = xls_path.parent / f"{xls_path.stem}.xlsx"
    if not xlsx_path.exists():
        raise FileNotFoundError(
            f".xls → .xlsx 변환 실패: {xls_path.name} (stderr: {result.stderr})"
        )

    logger.info(".xls → .xlsx 변환 완료: %s", xlsx_path.name)
    return xlsx_path


def _forward_fill_merged_cells(ws) -> dict[tuple[int, int], str]:
    """병합 셀의 좌상단 값을 병합 영역 전체에 복사 (forward-fill).

    Returns:
        {(row, col): value} — 병합 영역의 하위 셀에 채워진 값
    """
    filled = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = ws.cell(merged_range.min_row, merged_range.min_col)
        value = str(top_left.value).strip() if top_left.value is not None else ""
        if not value:
            continue
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if (row, col) != (merged_range.min_row, merged_range.min_col):
                    filled[(row, col)] = value
    return filled


def _get_cell_value(ws, row: int, col: int, filled: dict) -> str:
    """셀 값을 가져오되, 병합 셀이면 forward-fill 값 사용."""
    if (row, col) in filled:
        return filled[(row, col)]
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return filled.get((row, col), "")
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def _get_raw_cell_value(ws, row: int, col: int) -> str:
    """forward-fill 없이 원본 셀 값만 읽기. MergedCell은 빈 값 반환."""
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return ""
    if cell.value is None:
        return ""
    return str(cell.value).strip()


_GROUP_KEYWORDS = {"구분", "분류", "부문", "부서", "영역", "카테고리", "항목", "분야"}


def _find_group_col(
    filled: dict, header_row: int, headers: list[str],
) -> int | None:
    """Forward-fill 데이터에서 그룹핑 열 탐지.

    우선순위:
      1. 헤더명에 '구분' 등 키워드가 포함된 열 (가장 확실)
      2. forward-fill 셀이 많고, 고유값이 적은 열 (반복이 많은 = 그룹핑 열)
    """
    # 1순위: 헤더 키워드 매칭
    for col_idx, h in enumerate(headers):
        if h and any(kw in h for kw in _GROUP_KEYWORDS):
            col = col_idx + 1  # 1-based
            # 실제로 forward-fill이 있는 열인지 확인
            has_fill = any(c == col and r > header_row for (r, c) in filled)
            if has_fill:
                return col

    # 2순위: forward-fill 빈도 + 고유값 비율 기반 점수
    col_values: dict[int, list[str]] = {}
    for (row, col) in filled:
        if row > header_row:
            val = filled[(row, col)]
            col_values.setdefault(col, []).append(val)

    if not col_values:
        return None

    best_col = None
    best_score = -1.0
    for col, values in col_values.items():
        count = len(values)
        n_unique = len(set(values))
        avg_len = sum(len(v) for v in values) / count
        # 좋은 그룹핑 열: 많이 반복(count↑), 고유값 적음(n_unique↓), 값이 짧음(avg_len↓)
        score = count / max(n_unique, 1) / max(avg_len, 1)
        if score > best_score:
            best_score = score
            best_col = col

    return best_col


def _parse_sheet(ws) -> list[tuple[str, dict]]:
    """단일 시트를 구조화 텍스트 청크 리스트로 변환.

    섹션(구분) 경계에서 청크를 분리하여 검색 정확도를 높임.

    Returns:
        [(chunk_text, {"sheet": str, "section": str}), ...] 청크 + 메타데이터
    """
    if ws.max_row is None or ws.max_row < 1:
        return []

    max_col = ws.max_column or 1

    # 1단계: 헤더 행 탐지 (forward-fill 없이 원본 셀 값 사용)
    # 병합 타이틀은 좌상단 1셀만 값 있고 나머지는 MergedCell(빈 값)
    # → 고유 값 1~2개로 자연스럽게 건너뜀
    # 실제 헤더 행은 다양한 열 이름(No, 구분, 팀장, 그룹장 등)이 있어 고유 값이 많음
    headers = []
    header_row = 0
    for row in range(1, min(ws.max_row + 1, 20)):
        row_values = [
            _get_raw_cell_value(ws, row, col)
            for col in range(1, max_col + 1)
        ]
        non_empty = [v for v in row_values if v]
        unique_values = set(non_empty)
        if len(unique_values) >= 3:
            headers = row_values
            header_row = row
            logger.info("헤더 행 탐지: row %d (%d 고유 값: %s)",
                        row, len(unique_values),
                        ", ".join(list(unique_values)[:8]))
            break

    if not headers:
        return []

    # 2단계: forward-fill (헤더 확정 후, 데이터 영역의 병합 셀 해제)
    filled = _forward_fill_merged_cells(ws)

    # 3단계: 그룹핑 열 탐지 (섹션 경계 분리용)
    group_col = _find_group_col(filled, header_row, headers)
    if group_col:
        group_header = (
            headers[group_col - 1]
            if group_col - 1 < len(headers) and headers[group_col - 1]
            else "구분"
        )
        logger.info("그룹핑 열 탐지: col %d (%s)", group_col, group_header)

    # 4단계: 데이터 행을 구조화 텍스트로 변환 (섹션 경계에서 청크 분리)
    sheet_title = ws.title or "Sheet"
    chunks: list[tuple[str, dict]] = []
    current_lines: list[str] = []
    current_len = 0
    current_group: str | None = None

    def _make_chunk_header(group_value: str | None) -> str:
        if group_value and group_col:
            return f"[시트: {sheet_title}] [{group_header}: {group_value}]\n\n"
        return f"[시트: {sheet_title}]\n\n"

    def _make_meta(group_value: str | None) -> dict:
        meta = {"sheet": sheet_title}
        if group_value:
            meta["section"] = group_value
        return meta

    chunk_header = _make_chunk_header(None)

    for row in range(header_row + 1, (ws.max_row or 0) + 1):
        # 그룹 값 확인 (섹션 경계 감지)
        if group_col:
            group_value = _get_cell_value(ws, row, group_col, filled)
        else:
            group_value = None

        # 섹션이 바뀌면 현재 청크를 저장하고 새 청크 시작
        if group_value and group_value != current_group and current_lines:
            chunk_text = chunk_header + "".join(current_lines)
            chunks.append((chunk_text.strip(), _make_meta(current_group)))
            current_lines = []
            current_len = 0

        if group_value and group_value != current_group:
            current_group = group_value
            chunk_header = _make_chunk_header(current_group)

        # 행 텍스트 생성
        parts = []
        for col in range(1, (ws.max_column or 1) + 1):
            value = _get_cell_value(ws, row, col, filled)
            if not value:
                continue
            header = headers[col - 1] if col - 1 < len(headers) and headers[col - 1] else f"열{col}"
            parts.append(f"{header}: {value}")

        if not parts:
            continue

        line_text = f"[행] {' | '.join(parts)}\n"

        # 같은 섹션 내에서도 MAX_CHUNK_CHARS 초과 시 분할
        if current_len + len(line_text) > MAX_CHUNK_CHARS and current_lines:
            chunk_text = chunk_header + "".join(current_lines)
            chunks.append((chunk_text.strip(), _make_meta(current_group)))
            current_lines = []
            current_len = 0

        current_lines.append(line_text)
        current_len += len(line_text)

    if current_lines:
        chunk_text = chunk_header + "".join(current_lines)
        chunks.append((chunk_text.strip(), _make_meta(current_group)))

    return chunks


def parse_excel(file_path: Path) -> tuple[list[str], list[dict]]:
    """Excel 파일을 시트별 구조화 텍스트 리스트로 변환.

    .xls 파일은 LibreOffice로 .xlsx 변환 후 파싱.

    Returns:
        (texts, chunk_metas)
        - texts: 청크 텍스트 리스트 (= Qdrant "페이지" 단위)
        - chunk_metas: 각 청크의 메타 정보 [{"sheet": str, "section": str}, ...]
    """
    temp_xlsx = None
    actual_path = file_path

    if file_path.suffix.lower() == ".xls":
        actual_path = _convert_xls_to_xlsx(file_path)
        temp_xlsx = actual_path

    try:
        # read_only=False: 병합 셀 정보 접근에 필요
        wb = load_workbook(str(actual_path), read_only=False, data_only=True)
        all_chunks: list[tuple[str, dict]] = []
        num_sheets = len(wb.sheetnames)

        for ws in wb.worksheets:
            chunks = _parse_sheet(ws)
            all_chunks.extend(chunks)

        wb.close()

        if not all_chunks:
            return [f"[파일: {file_path.name}] (내용 없음)"], [{}]

        texts = [text for text, _ in all_chunks]
        metas = [meta for _, meta in all_chunks]

        logger.info(
            "Excel 파싱 완료: %s (%d 시트 → %d 청크)",
            file_path.name, num_sheets, len(texts),
        )
        return texts, metas

    finally:
        if temp_xlsx and temp_xlsx.exists():
            temp_xlsx.unlink()
